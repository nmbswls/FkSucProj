# 通缉表：维护 wanted.xlsx（行为表增加单次上限、新增守卫档位 sheet），并注册 TbWantedGuardSpawnTier。
# 在 Config 目录执行: python gen_wanted_tables.py 后运行 gen.bat（或 dotnet Luban）。
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DATAS = ROOT / "Datas"
WANTED_XLSX = DATAS / "wanted.xlsx"
TABLES_XLSX = DATAS / "__tables__.xlsx"


def write_wanted_workbook():
    wb = openpyxl.Workbook()

    # --- wanted_level_info ---
    ws_l = wb.active
    ws_l.title = "wanted_level_info"
    ws_l.append(["##var", "level", "need_val"])
    ws_l.append(["##type", "int", "int"])
    ws_l.append(["##group", "c,s", "c,s"])
    ws_l.append(["##", "id", "累计通缉阈值(×1000 后与实际 CurrentWantedVal 比较)"])
    for lv, nv in [(0, 0), (1, 20), (2, 40), (3, 60), (4, 80), (5, 100)]:
        ws_l.append([None, lv, nv])

    # --- wanted_behave_info ---
    ws_b = wb.create_sheet("wanted_behave_info")
    ws_b.append(
        [
            "##var",
            "BehaveType",
            "add_wanted",
            "trigger_range",
            "max_add_once",
        ]
    )
    ws_b.append(["##type", "EWantedBehaveType", "int", "float", "int"])
    ws_b.append(["##group", "c,s", "c,s", "c,s", "c,s"])
    ws_b.append(
        [
            "##",
            "id",
            "每次基础增量(传入 AddWantedVal 的逻辑量，会再×1000)",
            "NPC 感知/视线相关占位",
            "单次叠加封顶(逻辑量，0=不限制)与 add_wanted 取较小",
        ]
    )
    ws_b.append([None, "StealSmall", 1, 3.0, 8])
    ws_b.append([None, "StealValuable", 3, 5.0, 15])
    ws_b.append([None, "AssaultCitizen", 2, 8.0, 12])

    # --- wanted_guard_spawn ---
    ws_g = wb.create_sheet("wanted_guard_spawn")
    ws_g.append(
        [
            "##var",
            "tier_id",
            "min_wanted_star_level",
            "guard_count",
            "npc_cfg_id",
            "spawn_radius_min",
            "spawn_radius_max",
            "cull_distance",
        ]
    )
    ws_g.append(["##type", "int", "int", "int", "string", "float", "float", "float"])
    ws_g.append(["##group", "c,s", "c,s", "c,s", "c,s", "c,s", "c,s", "c,s"])
    ws_g.append(
        [
            "##",
            "id",
            "通缉星级(见 wanted_level_info)",
            "守卫数",
            "NPC",
            "环内径",
            "环外径",
            "超出距离销毁",
        ]
    )
    rows = [
        (1, 0, 0, "default_guard_01", 8.0, 14.0, 40.0),
        (2, 2, 1, "default_guard_01", 6.0, 12.0, 36.0),
        (3, 3, 2, "default_guard_01", 5.0, 11.0, 34.0),
        (4, 4, 3, "default_guard_01", 4.5, 10.0, 32.0),
        (5, 5, 4, "default_guard_01", 4.0, 9.0, 30.0),
    ]
    for r in rows:
        ws_g.append([None, *r])

    wb.save(WANTED_XLSX)


def patch_tables():
    wb = openpyxl.load_workbook(TABLES_XLSX)
    ws = wb.active
    full = "demo.TbWantedGuardSpawnTier"
    target_row = None
    for r in range(4, ws.max_row + 1):
        cell = ws.cell(r, 2).value
        if cell is not None and str(cell) == full:
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(target_row, 1, None)
    ws.cell(target_row, 2, full)
    ws.cell(target_row, 3, "demo.WantedGuardSpawnTier")
    ws.cell(target_row, 4, True)
    ws.cell(target_row, 5, "wanted_guard_spawn@wanted.xlsx")
    ws.cell(target_row, 6, None)
    ws.cell(target_row, 7, None)
    ws.cell(target_row, 8, None)
    wb.save(TABLES_XLSX)


if __name__ == "__main__":
    DATAS.mkdir(parents=True, exist_ok=True)
    write_wanted_workbook()
    patch_tables()
    legacy = DATAS.parent / "output" / "demo_tbwantedguardspawntier.json"
    if legacy.exists():
        legacy.unlink()
    print("ok wanted xlsx + tables patched; run gen.bat for Luban")
