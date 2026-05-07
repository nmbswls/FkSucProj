# 天赋树：生成 talent.xlsx 并在 __tables__.xlsx 注册 TbTalentNode（read_schema_from_file）。
# 在 Config 目录执行: python gen_talent_tables.py 后运行 gen.bat。
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DATAS = ROOT / "Datas"
OUT_XLSX = DATAS / "talent.xlsx"
TABLES_XLSX = DATAS / "__tables__.xlsx"

SHEET = "talent"

COND_NONE = '0,0,0,0,0,"",""'


def write_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET

    ws.append(
        [
            "##var",
            "node_id",
            "display_name",
            "prereq_node_ids",
            "unlock_conds",
            "unlock_costs",
            "stat_bonuses",
            "passive_skill_id",
        ]
    )
    ws.append(
        [
            "##type",
            "int",
            "string",
            "(list#sep=;),int",
            "(list#sep=;),CommonCheckCond",
            "(list#sep=|),TalentUnlockCost",
            "(list#sep=|),TalentStatBonus",
            "string",
        ]
    )
    ws.append(
        [
            "##group",
            None,
            "c,s",
            "c",
            "c",
            "c",
            "c",
            "c",
            "c",
        ]
    )
    ws.append(
        [
            "##",
            "node_id",
            "display_name",
            "prereq_node_ids",
            "unlock_conds",
            "unlock_costs",
            "stat_bonuses",
            "passive_skill_id",
        ]
    )

    # EYCAttribute: InnerCharm=2, StaticCharm=3, InnerArm=4, StaticArm=5, SecretSlot=1
    rows = [
        (None, 1, "天分-魅力根基", "", COND_NONE, "", "2,1200", ""),
        (None, 2, "天分-护甲理解", "1", COND_NONE, "wood,2", "4,800", ""),
        (None, 3, "天分-意志魅力", "1", COND_NONE, "stick,1", "3,600", ""),
        (None, 4, "天分-双防雏形", "2", COND_NONE, "wood,1|stick,1", "5,700|4,400", ""),
        (None, 5, "天分-统合", "2;3", COND_NONE, "gold,5", "2,500|5,300", ""),
    ]
    for r in rows:
        ws.append(list(r))

    wb.save(OUT_XLSX)


def patch_tables():
    wb = openpyxl.load_workbook(TABLES_XLSX)
    ws = wb.active
    full = "demo.TbTalentNode"
    target_row = None
    for r in range(4, ws.max_row + 1):
        cell = ws.cell(r, 2).value
        if cell is not None and str(cell) == full:
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(target_row, 1, None)
    ws.cell(target_row, 2, full)
    ws.cell(target_row, 3, "demo.TalentNode")
    ws.cell(target_row, 4, True)
    ws.cell(target_row, 5, f"{SHEET}@talent.xlsx")
    ws.cell(target_row, 6, None)
    ws.cell(target_row, 7, None)
    ws.cell(target_row, 8, None)
    wb.save(TABLES_XLSX)


if __name__ == "__main__":
    DATAS.mkdir(parents=True, exist_ok=True)
    write_workbook()
    patch_tables()
    legacy = DATAS / "demo_tbtalentnode.json"
    if legacy.exists():
        legacy.unlink()
    print("ok")
