# Generate item.xlsx + item_use.xlsx and register TbItemUse. Run from repo root or Config/.
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent
DATAS = ROOT / "Datas"

ITEM_COLS = [
    ("item_id", "string"),
    ("display_name", "string"),
    ("item_type", "EItemType"),
    ("stack_type", "EItemStackType"),
    ("stack_count", "int"),
    ("max_stack_inventory", "int"),
    ("max_stack_shop", "int"),
    ("max_stack_loot", "int"),
    ("sprite_name", "string"),
    ("can_drop", "bool"),
    ("rare_tier", "int"),
    ("reveal_effect_type", "EItemRevealEffectType"),
    ("reveal_p1", "long"),
    ("reveal_p2", "long"),
    ("auto_destroy", "bool"),
    ("auto_destroy_time", "float"),
    ("special_buff_id", "string"),
    ("special_buff_interval", "float"),
    ("auto_pick", "bool"),
    ("is_auto_use", "bool"),
]


def write_item_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "item"

    ws.append(["##var"] + [c[0] for c in ITEM_COLS])
    ws.append(["##type"] + [c[1] for c in ITEM_COLS])
    ws.append(["##group"] + ["c,s"] * len(ITEM_COLS))
    ws.append(["##"] + [c[0] for c in ITEM_COLS])

    def row(
        item_id,
        display_name,
        item_type,
        stack_type,
        stack_count=0,
        ms_inv=0,
        ms_shop=0,
        ms_loot=0,
        sprite=None,
        can_drop=True,
        rare_tier=0,
        reveal_t="None",
        rp1=0,
        rp2=0,
        auto_des=False,
        auto_des_t=0.0,
        buff_id="",
        buff_iv=0.0,
        auto_pick=False,
        auto_use=False,
    ):
        sp = sprite if sprite is not None else item_id
        return [
            None,
            item_id,
            display_name,
            item_type,
            stack_type,
            stack_count,
            ms_inv,
            ms_shop,
            ms_loot,
            sp,
            can_drop,
            rare_tier,
            reveal_t,
            rp1,
            rp2,
            auto_des,
            auto_des_t,
            buff_id,
            buff_iv,
            auto_pick,
            auto_use,
        ]

    rows = [
        row("1", "1", "Normal", "Size1"),
        row("small_stone", "small_stone", "Normal", "Size1"),
        row("stick", "stick", "Normal", "Size2"),
        row("wood", "wood", "Normal", "Size3"),
        row("banana", "banana", "Normal", "Size2", reveal_t="AddGcVal", rp1=5000),
        row("qiezi", "qiezi", "Normal", "Size2", reveal_t="AddGcVal", rp1=5000),
        row("bangbangtang", "bangbangtang", "Normal", "Size2"),
        row("flower_01", "flower_01", "Normal", "Size2"),
        row("flower_02", "flower_02", "Normal", "Size2"),
        row("flower_03", "flower_03", "Normal", "Size2"),
        row("berry", "berry", "Normal", "Size1"),
        row("gold", "gold", "Currency", "NoLimit"),
        row("j", "j", "Currency", "NoLimit"),
        row("chanzi", "chanzi", "Equip", "NoStack"),
        row("key_a1_001", "key_a1_001", "Normal", "NoStack", sprite="key_a1_01"),
        row("key_a1_002", "key_a1_002", "Normal", "NoStack", sprite="key_a1_02"),
        row(
            "insertion_maoqiu",
            "insertion_maoqiu",
            "Insertion",
            "NoStack",
            can_drop=False,
            auto_des=True,
            auto_des_t=30.0,
            buff_id="insertion_debuff_small",
            buff_iv=5.0,
        ),
        row(
            "j_drop_small",
            "j_drop_small",
            "Normal",
            "NoStack",
            auto_pick=True,
            auto_use=True,
        ),
        row("evil_scroll_01", "evil_scroll_01", "Normal", "Size1"),
    ]

    for r in rows:
        ws.append(r)

    wb.save(DATAS / "item.xlsx")


USE_COLS = [
    ("id", "int"),
    ("item_id", "string"),
    ("slot", "int"),
    ("usable", "bool"),
    ("cost_on_use", "bool"),
    ("use_cd", "float"),
    ("use_time", "float"),
    ("use_type", "EItemUseType"),
    ("p1", "long"),
    ("p2", "long"),
    ("p3", "long"),
    ("p4", "long"),
    ("s1", "string"),
    ("s2", "string"),
]


def write_item_use_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "item_use"

    ws.append(["##var"] + [c[0] for c in USE_COLS])
    ws.append(["##type"] + [c[1] for c in USE_COLS])
    ws.append(["##group"] + ["c,s"] * len(USE_COLS))
    ws.append(["##"] + [c[0] for c in USE_COLS])

    # id, item_id, slot, usable, cost_on_use, use_cd, use_time, use_type, p1..p4, s1, s2
    use_rows = [
        (1, "banana", 1, True, True, 10.0, 1.5, "None", 0, 0, 0, 0, None, None),
        (
            2,
            "j_drop_small",
            1,
            True,
            True,
            0.0,
            0.0,
            "GiveDrop",
            100,
            0,
            0,
            0,
            None,
            None,
        ),
        (
            3,
            "evil_scroll_01",
            1,
            True,
            True,
            5.0,
            0.0,
            "UseAbility",
            0,
            0,
            0,
            0,
            "queen_pull_all",
            None,
        ),
    ]

    for u in use_rows:
        ws.append([None] + list(u))

    wb.save(DATAS / "item_use.xlsx")


def patch_tables():
    wb = openpyxl.load_workbook(DATAS / "__tables__.xlsx")
    ws = wb.active
    # skip if already registered
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 5).value
        if v and "item_use" in str(v):
            wb.save(DATAS / "__tables__.xlsx")
            return

    r = ws.max_row + 1
    ws.cell(r, 1, None)
    ws.cell(r, 2, "demo.TbItemUse")
    ws.cell(r, 3, "demo.ItemUse")
    ws.cell(r, 4, True)
    ws.cell(r, 5, "item_use@item_use.xlsx")
    wb.save(DATAS / "__tables__.xlsx")


if __name__ == "__main__":
    write_item_xlsx()
    write_item_use_xlsx()
    patch_tables()
    print("ok")
