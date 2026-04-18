# 入梦入口点位：生成 dream_infiltration.xlsx 并在 __tables__.xlsx 注册 TbDreamInfiltrationSpot。
# 在 Config 目录执行: python gen_dream_infiltration_tables.py 后运行 gen.bat。
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DATAS = ROOT / "Datas"
OUT_XLSX = DATAS / "dream_infiltration.xlsx"
TABLES_XLSX = DATAS / "__tables__.xlsx"

SHEET = "dream_infiltration"

# unlock_conds: (list#sep=;),CommonCheckCond
# theme_ids / theme_display_names / theme_weight_values：三列平行 (list#sep=;)，与 drop.xlsx 等表一致


def write_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET

    ws.append(
        [
            "##var",
            "spot_id",
            "display_name",
            "anchor_x",
            "anchor_y",
            "unlock_conds",
            "theme_ids",
            "theme_display_names",
            "theme_weight_values",
        ]
    )
    ws.append(
        [
            "##type",
            "string",
            "string",
            "float",
            "float",
            "(list#sep=;),CommonCheckCond",
            "(list#sep=;),string",
            "(list#sep=;),string",
            "(list#sep=;),int",
        ]
    )
    ws.append(
        [
            "##group",
            None,
            "c,s",
            "c",
            "c",
            "c",
            "c",
            "c",
            "c",
        ]
    )
    ws.append(
        [
            "##",
            "spot_id",
            "display_name",
            "anchor_x",
            "anchor_y",
            "unlock_conds",
            "theme_ids",
            "theme_display_names",
            "theme_weight_values",
        ]
    )

    tid = "ruins;garden;maze"
    tnm = "废墟回响;花园低语;迷宫心象"
    tw = "10;10;10"

    cond_none = '0,0,0,0,0,"",""'
    cond_always_fail = '6,0,0,0,0,"",""'

    rows = [
        (None, "north", "北门潜入口", 0.22, 0.62, cond_none, tid, tnm, tw),
        (None, "east", "东侧裂隙", 0.72, 0.48, cond_none, tid, tnm, tw),
        (
            None,
            "locked_demo",
            "需条件（演示 AlwaysFail）",
            0.48,
            0.28,
            cond_always_fail,
            tid,
            tnm,
            tw,
        ),
    ]
    for r in rows:
        ws.append(list(r))

    wb.save(OUT_XLSX)


def patch_tables():
    wb = openpyxl.load_workbook(TABLES_XLSX)
    ws = wb.active
    full = "demo.TbDreamInfiltrationSpot"
    target_row = None
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 2).value) == full:
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(target_row, 1, None)
    ws.cell(target_row, 2, full)
    ws.cell(target_row, 3, "demo.DreamInfiltrationSpot")
    ws.cell(target_row, 4, False)
    ws.cell(target_row, 5, f"{SHEET}@dream_infiltration.xlsx")
    ws.cell(target_row, 6, None)
    ws.cell(target_row, 7, None)
    ws.cell(target_row, 8, None)
    wb.save(TABLES_XLSX)


if __name__ == "__main__":
    DATAS.mkdir(parents=True, exist_ok=True)
    write_workbook()
    patch_tables()
    legacy = DATAS / "demo_tbdreaminfiltrationspot.json"
    if legacy.exists():
        legacy.unlink()
    print("ok")
