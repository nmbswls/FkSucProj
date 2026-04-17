# 生成 map.xlsx 中大地图路由相关 sheet，并注册到 __tables__.xlsx。在 Config 目录执行: python gen_world_map_tables.py
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DATAS = ROOT / "Datas"
MAP_XLSX = DATAS / "map.xlsx"
TABLES_XLSX = DATAS / "__tables__.xlsx"


def _append_meta(ws, var_row, type_row, group_row, note_row):
    ws.append(["##var"] + var_row)
    ws.append(["##type"] + type_row)
    ws.append(["##group"] + group_row)
    ws.append(["##"] + note_row)


def write_world_map_sheets():
    wb = openpyxl.load_workbook(MAP_XLSX)

    # 移除旧 sheet（若存在）以便整表重建
    for name in ("world_map_area", "world_map_room_rule", "world_map_settings"):
        if name in wb.sheetnames:
            del wb[name]

    ws_area = wb.create_sheet("world_map_area")
    _append_meta(
        ws_area,
        ["id", "map_texture_resource_path", "world_min_x", "world_min_y", "world_max_x", "world_max_y"],
        ["string", "string", "float", "float", "float", "float"],
        [None, "c,s", "c", "c", "c", "c"],
        ["区域 id（与 MapAreaInfo.id / MapName 一致）", "Resources 路径无扩展名", "min", "min", "max", "max"],
    )
    rows_area = [
        ("base_01", "WorldMap/fake_map_base_01", -50, -50, 50, 50),
        ("village_01", "WorldMap/fake_map_village_01", -50, -50, 50, 50),
        ("game_init", "WorldMap/fake_map_game_init", -50, -50, 50, 50),
    ]
    for r in rows_area:
        ws_area.append([None, *r])

    ws_rule = wb.create_sheet("world_map_room_rule")
    _append_meta(
        ws_rule,
        [
            "id",
            "area_id",
            "room_id",
            "rule_priority",
            "behavior",
            "alternate_map_texture_resource_path",
            "use_separate_bounds",
            "alternate_world_min_x",
            "alternate_world_min_y",
            "alternate_world_max_x",
            "alternate_world_max_y",
        ],
        [
            "int",
            "string",
            "string",
            "int",
            "EWorldMapRoomBehavior",
            "string",
            "bool",
            "float",
            "float",
            "float",
            "float",
        ],
        [None, "c,s", "c", "c", "c", "c", "c", "c", "c", "c", "c"],
        [
            "规则 id",
            "所属区域",
            "房间 id，* 为默认",
            "同区多条时大者优先",
            "行为",
            "换图时纹理路径",
            "是否用独立边界",
            "alt min/max",
            "",
            "",
            "",
        ],
    )
    # 无示例行：按需添加；Luban 接受空表

    ws_set = wb.create_sheet("world_map_settings")
    _append_meta(
        ws_set,
        [
            "allow_open_when_area_unknown",
            "fallback_map_texture_resource_path",
            "fallback_world_min_x",
            "fallback_world_min_y",
            "fallback_world_max_x",
            "fallback_world_max_y",
            "global_npc_boss_landmark_cfg_ids",
            "global_interact_landmark_cfg_ids",
        ],
        ["bool", "string", "float", "float", "float", "float", "string", "string"],
        [None, "c,s", "c", "c", "c", "c", "c", "c"],
        [
            "未知区域是否允许打开",
            "fallback 纹理 Resources 路径",
            "fallback 边界",
            "",
            "",
            "",
            "逗号分隔 cfgId",
            "逗号分隔 cfgId",
        ],
    )
    # mode=one：仅一行数据
    ws_set.append(
        [
            None,
            True,
            "WorldMap/fake_map_fallback",
            -40,
            -40,
            40,
            40,
            "",
            "",
        ]
    )

    wb.save(MAP_XLSX)


def patch_tables():
    wb = openpyxl.load_workbook(TABLES_XLSX)
    ws = wb.active

    want = [
        ("demo.TbWorldMapArea", "demo.WorldMapArea", True, "world_map_area@map.xlsx", None, None),
        ("demo.TbWorldMapRoomRule", "demo.WorldMapRoomRule", True, "world_map_room_rule@map.xlsx", None, "list"),
        ("demo.TbWorldMapSettings", "demo.WorldMapSettings", True, "world_map_settings@map.xlsx", None, "one"),
    ]

    existing = set()
    for r in range(4, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v:
            existing.add(str(v))

    for full_name, value_type, read_schema, inp, index, mode in want:
        if full_name in existing:
            for r in range(4, ws.max_row + 1):
                if str(ws.cell(r, 2).value) == full_name:
                    ws.cell(r, 3, value_type)
                    ws.cell(r, 4, read_schema)
                    ws.cell(r, 5, inp)
                    ws.cell(r, 6, index)
                    ws.cell(r, 7, mode)
            continue
        nr = ws.max_row + 1
        ws.cell(nr, 1, None)
        ws.cell(nr, 2, full_name)
        ws.cell(nr, 3, value_type)
        ws.cell(nr, 4, read_schema)
        ws.cell(nr, 5, inp)
        ws.cell(nr, 6, index)
        ws.cell(nr, 7, mode)

    wb.save(TABLES_XLSX)


if __name__ == "__main__":
    write_world_map_sheets()
    patch_tables()
    print("ok")
