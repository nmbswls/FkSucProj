# 垂钓：单文件 fishing_spot_config.xlsx（两 sheet），两个 Luban 表（见 __tables__.xlsx）。
#   fishing_spot      — 钓点：id、presentation_prefab_path（Resources 下预制体 key）、显示名、容量、补鱼周期
#   fishing_spot_fish — 每钓点每鱼种一行：spot_id 关联主表 id，unlock_conds，weight
# 在 Config 目录: python gen_fishing_spot_tables.py 后 gen.bat。（若 xlsx 被 Excel 占用，先关闭再运行）
from __future__ import annotations

import os
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DATAS = ROOT / "Datas"
# 使用独立文件名，避免 fishing_spot.xlsx 被 Excel 占用时无法替换
OUT_XLSX = DATAS / "fishing_spot_config.xlsx"
TABLES_XLSX = DATAS / "__tables__.xlsx"

SHEET_SPOT = "fishing_spot"
SHEET_FISH = "fishing_spot_fish"

SPOT_COLS = [
    ("id", "string", "钓点配置 id，与地图 InitInfo.CfgId 一致"),
    (
        "presentation_prefab_path",
        "string",
        "表现预制体 Resources 相对路径，如 Prefab/Presentations/FishingSpot/xxx（与 PresentationFactory 加载 key 一致）",
    ),
    ("display_name", "string", "显示名"),
    ("capacity", "int", "可钓次数上限"),
    ("restock_every_n_days", "int", "每多少个结算日补满容量"),
]

# 主键 line_id；spot_id 与 fishing_spot.id 对应
FISH_COLS = [
    ("line_id", "int", "行主键，全表唯一"),
    ("spot_id", "string", "关联 fishing_spot.id"),
    ("fish_item_id", "string", "钓到的道具/鱼 item id"),
    (
        "unlock_conds",
        "(list#sep=;),CommonCheckCond",
        "解锁条件列表，空单元或单条 None 条件表示无限制（与 dream unlock_conds 填法一致）",
    ),
    ("weight", "int", "权重（仅对已解锁条目参与加权）"),
]


def _append_sheet_meta(ws, cols):
    ws.append(["##var"] + [c[0] for c in cols])
    ws.append(["##type"] + [c[1] for c in cols])
    ws.append(["##group"] + ["c,s"] * len(cols))
    ws.append(["##"] + [c[2] for c in cols])


def _atomic_save_workbook(wb: openpyxl.Workbook, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(dest.parent))
    os.close(fd)
    tmp_path = Path(tmp)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, dest)
    except Exception:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
        raise


def write_workbook():
    wb = openpyxl.Workbook()
    # --- 主表 ---
    ws = wb.active
    ws.title = SHEET_SPOT
    _append_sheet_meta(ws, SPOT_COLS)

    spots = [
        (
            "fish_spot_base_01",
            "Prefab/Presentations/FishingSpot/fish_spot_base_01",
            "Demo Pond",
            5,
            3,
        ),
        (
            "fish_spot_base_02",
            "Prefab/Presentations/FishingSpot/fish_spot_base_02",
            "Demo Pond",
            5,
            3,
        ),
        (
            "fish_spot_base_03",
            "Prefab/Presentations/FishingSpot/fish_spot_base_03",
            "Demo Pond",
            5,
            3,
        ),
        (
            "fish_spot_base_04",
            "Prefab/Presentations/FishingSpot/fish_spot_base_04",
            "Demo Pond",
            5,
            3,
        ),
    ]
    for row in spots:
        ws.append([None, *row])

    # --- 子表：按 spot_id 关联 ---
    wf = wb.create_sheet(SHEET_FISH)
    _append_sheet_meta(wf, FISH_COLS)
    cond_none = '0,0,0,0,0,"",""'
    line_id = 1
    for sid, *_rest in spots:
        wf.append([None, line_id, sid, "banana", cond_none, 40])
        line_id += 1
        wf.append([None, line_id, sid, "qiezi", cond_none, 60])
        line_id += 1

    _atomic_save_workbook(wb, OUT_XLSX)


def _set_table_row(ws, full_name: str, value_type: str, read_schema: bool, inp: str):
    target_row = None
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, 2).value) == full_name:
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(target_row, 1, None)
    ws.cell(target_row, 2, full_name)
    ws.cell(target_row, 3, value_type)
    ws.cell(target_row, 4, read_schema)
    ws.cell(target_row, 5, inp)
    ws.cell(target_row, 6, None)
    ws.cell(target_row, 7, None)
    ws.cell(target_row, 8, None)


def patch_tables():
    wb = openpyxl.load_workbook(TABLES_XLSX)
    ws = wb.active
    _set_table_row(
        ws,
        "demo.TbFishingSpot",
        "demo.FishingSpot",
        True,
        f"{SHEET_SPOT}@{OUT_XLSX.name}",
    )
    _set_table_row(
        ws,
        "demo.TbFishingSpotFish",
        "demo.FishingSpotFish",
        True,
        f"{SHEET_FISH}@{OUT_XLSX.name}",
    )
    wb.save(TABLES_XLSX)


if __name__ == "__main__":
    DATAS.mkdir(parents=True, exist_ok=True)
    write_workbook()
    patch_tables()
    print("ok")
