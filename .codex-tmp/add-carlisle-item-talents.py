from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\UProjects\FkSucProj\FkSucProj\Config\Datas")


def copy_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def write_row(ws, row, values, style_row):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row, col, value)
        copy_style(ws.cell(style_row, min(col, ws.max_column)), cell)


talent_path = ROOT / "talent.xlsx"
wb = load_workbook(talent_path)
nodes = wb["talent"]

if nodes.cell(1, 7).value != "description":
    nodes.cell(1, 7, "description")
    nodes.cell(2, 7, "string")
    nodes.cell(3, 7, "c")
    nodes.cell(4, 7, "节点描述")
    for row in range(1, nodes.max_row + 1):
        copy_style(nodes.cell(row, 6), nodes.cell(row, 7))

node_rows = [
    (None, "carlisle_tree", 104, "旧囊加层", 1, "default", "真正耐用的行囊不靠外表。卡莱尔教你用捡来的布片和绳头，在旧背包里再缝出一层藏物夹袋。"),
    (None, "carlisle_tree", 105, "负肩成习", 1, "default", "旅途中没有人替你分担重量。卡莱尔用旧合页和铁钉改出一套负重带，教你让肩背逐渐习惯长路。"),
    (None, "carlisle_tree", 106, "硬地筋骨", 1, "default", "没有训练场的流浪者，就拿路边和废墟当老师。用铁锭与亚麻布做成简陋负具，反复锤炼肉体。"),
]
for values in node_rows:
    row = next((r for r in range(5, nodes.max_row + 1) if nodes.cell(r, 3).value == values[2]), None)
    write_row(nodes, row or nodes.max_row + 1, values, 5)

levels = wb["talent_level"]
level_rows = [
    (None, 104, 1, "旧囊加层", "", "", "loot_cloth_scrap,4|loot_rope_end,2", "13,1", ""),
    (None, 105, 1, "负肩成习", "", "", "loot_bent_hinge,2|loot_rusty_nail,4", "21,500", ""),
    (None, 106, 1, "硬地筋骨", "", "", "mat_iron_ingot,1|mat_linen_cloth,2", "20,500", ""),
]
for values in level_rows:
    row = next((
        r for r in range(5, levels.max_row + 1)
        if levels.cell(r, 2).value == values[1] and levels.cell(r, 3).value == values[2]
    ), None)
    write_row(levels, row or levels.max_row + 1, values, 5)

wb.save(talent_path)


yc_path = ROOT / "yc_attribute.xlsx"
yc_wb = load_workbook(yc_path)
yc = yc_wb.active
existing = {yc.cell(row, 8).value for row in range(4, yc.max_row + 1)}
insert_at = next((row for row in range(4, yc.max_row + 1) if yc.cell(row, 10).value == 100), yc.max_row + 1)
new_attrs = [
    ("PhysicalPower", 20, "肉体强度"),
    ("PhysicalResist", 21, "肉体耐受"),
]
for name, value, comment in new_attrs:
    if name in existing:
        continue
    yc.insert_rows(insert_at)
    for col in range(1, yc.max_column + 1):
        copy_style(yc.cell(insert_at + 1, col), yc.cell(insert_at, col))
    yc.cell(insert_at, 8, name)
    yc.cell(insert_at, 10, value)
    yc.cell(insert_at, 11, comment)
    insert_at += 1

yc_wb.save(yc_path)
